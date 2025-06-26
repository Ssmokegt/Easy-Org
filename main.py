import tkinter as tk
from tkinter import ttk, messagebox, filedialog, colorchooser
from PIL import Image, ImageDraw, ImageFont, ImageTk
import os
import json
import csv
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import ImageReader
try:
    import openpyxl
except ImportError:
    openpyxl = None
import json
import csv
import platform

class Empleado:
    def __init__(self, nombre, puesto, departamento, supervisor, color):
        self.nombre = nombre
        self.puesto = puesto
        self.departamento = departamento
        self.supervisor = supervisor
        self.color = color

class OrgApp:
    def add_tooltip(self, widget, text):
        tooltip = tk.Toplevel(widget, bg='#333333', padx=6, pady=3)
        tooltip.wm_overrideredirect(True)
        tooltip.withdraw()
        label = tk.Label(tooltip, text=text, bg='#333333', fg='white', font=('Arial', 9))
        label.pack()
        def enter(event):
            x = widget.winfo_rootx() + 30
            y = widget.winfo_rooty() + 20
            tooltip.wm_geometry(f'+{x}+{y}')
            tooltip.deiconify()
        def leave(event):
            tooltip.withdraw()
        widget.bind('<Enter>', enter)
        widget.bind('<Leave>', leave)

    def mostrar_notificacion(self, mensaje, color='#2e7d32'):
        self.notif_label.config(text=mensaje, bg=color)
        self.notif_label.grid()
        self.root.after(2000, self.notif_label.grid_remove)

    def __init__(self, root):
        self.root = root
        self.root.title("Gestor de Organigrama Empresarial")
        self.empleados = []
        # Tema moderno
        style = ttk.Style()
        style.theme_use('clam')
        self.setup_ui()
        # Notificación temporal
        self.notif_label = tk.Label(self.root, text='', fg='white', bg='#2e7d32', font=('Arial', 11, 'bold'))
        self.notif_label.grid(row=99, column=0, columnspan=2, sticky='ew', pady=2)
        self.notif_label.grid_remove()
        # Cargar automáticamente empleados.json si existe
        self.cargar_automatica_empleados()

    def cargar_automatica_empleados(self):
        ruta = os.path.join(os.path.dirname(os.path.abspath(__file__)), "empleados.json")
        if os.path.exists(ruta):
            try:
                with open(ruta, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                self.empleados = [Empleado(emp["nombre"], emp["puesto"], emp["departamento"], emp["supervisor"], emp["color"]) for emp in data]
                self.actualizar_lista()
                self.actualizar_supervisores()
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo cargar empleados.json automáticamente:\n{e}")

    def setup_ui(self):
        # Frame de formulario
        form_frame = ttk.LabelFrame(self.root, text="Agregar/Editar Empleado")
        form_frame.grid(row=0, column=0, padx=14, pady=12, sticky="nsew")

        # Campos
        ttk.Label(form_frame, text="Nombre:").grid(row=0, column=0, sticky="e", pady=4, padx=2)
        self.nombre_var = tk.StringVar()
        nombre_entry = ttk.Entry(form_frame, textvariable=self.nombre_var)
        nombre_entry.grid(row=0, column=1, pady=4, sticky='ew')

        ttk.Label(form_frame, text="Puesto:").grid(row=1, column=0, sticky="e", pady=4, padx=2)
        self.puesto_var = tk.StringVar()
        puesto_entry = ttk.Entry(form_frame, textvariable=self.puesto_var)
        puesto_entry.grid(row=1, column=1, pady=4, sticky='ew')

        ttk.Label(form_frame, text="Departamento:").grid(row=2, column=0, sticky="e", pady=4, padx=2)
        self.depto_var = tk.StringVar()
        depto_entry = ttk.Entry(form_frame, textvariable=self.depto_var)
        depto_entry.grid(row=2, column=1, pady=4, sticky='ew')

        ttk.Label(form_frame, text="Supervisor:").grid(row=3, column=0, sticky="e", pady=4, padx=2)
        self.supervisor_var = tk.StringVar()
        self.supervisor_combo = ttk.Combobox(form_frame, textvariable=self.supervisor_var, values=[], state="readonly")
        self.supervisor_combo.grid(row=3, column=1, pady=4, sticky='ew')

        ttk.Label(form_frame, text="Color:").grid(row=4, column=0, sticky="e", pady=4, padx=2)
        self.color_var = tk.StringVar(value="#ffffff")
        color_btn = ttk.Button(form_frame, text="Seleccionar color", command=self.elegir_color)
        color_btn.grid(row=4, column=1, sticky="w", pady=4)

        # Botones
        add_btn = ttk.Button(form_frame, text="Agregar/Actualizar", command=self.agregar_empleado)
        add_btn.grid(row=5, column=0, columnspan=2, pady=8)

        # Tooltips para campos principales
        self.add_tooltip(nombre_entry, "Nombre completo del empleado")
        self.add_tooltip(puesto_entry, "Puesto que ocupa el empleado")
        self.add_tooltip(depto_entry, "Departamento al que pertenece")
        self.add_tooltip(self.supervisor_combo, "Seleccione el supervisor de este empleado")
        self.add_tooltip(color_btn, "Color del cuadro del empleado en el organigrama")
        self.add_tooltip(add_btn, "Agregar o actualizar empleado")

        # Lista de empleados
        lista_frame = ttk.LabelFrame(self.root, text="Empleados")
        lista_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        self.lista = tk.Listbox(lista_frame, height=12, width=52, selectbackground='#90caf9', selectforeground='black', activestyle='none')
        self.lista.pack(side="left", fill="both", expand=True, padx=2, pady=2)
        self.lista.bind('<<ListboxSelect>>', self.seleccionar_empleado)
        self.lista.bind('<Double-Button-1>', self.editar_empleado_rapido)
        delete_btn = ttk.Button(lista_frame, text="Eliminar", command=self.eliminar_empleado)
        delete_btn.pack(side="right", padx=8, pady=8)
        self.add_tooltip(self.lista, "Lista de empleados. Doble clic para editar. Arrastra para cambiar supervisor.")
        self.add_tooltip(delete_btn, "Eliminar empleado seleccionado")

        # Botón para generar organigrama
        generar_btn = ttk.Button(self.root, text="Generar Organigrama", command=self.generar_organigrama)
        generar_btn.grid(row=2, column=0, pady=10)

        # Botones de exportación
        btn_frame = ttk.Frame(self.root)
        btn_frame.grid(row=3, column=0, padx=10, pady=5)
        self.exportar_png_btn = ttk.Button(btn_frame, text="Exportar como PNG", command=self.exportar_png, state="disabled")
        self.exportar_png_btn.grid(row=0, column=0, padx=5, pady=2)
        self.add_tooltip(self.exportar_png_btn, "Exportar el organigrama como imagen PNG")
        self.exportar_pdf_btn = ttk.Button(btn_frame, text="Exportar a PDF", command=self.exportar_pdf)
        self.exportar_pdf_btn.grid(row=0, column=1, padx=5, pady=2)
        self.add_tooltip(self.exportar_pdf_btn, "Exporta el organigrama como PDF")
        self.exportar_pdf_btn.config(state="disabled")

        self.exportar_csv_btn = ttk.Button(btn_frame, text="Exportar a CSV", command=self.exportar_csv)
        self.exportar_csv_btn.grid(row=0, column=2, padx=5, pady=2)
        self.add_tooltip(self.exportar_csv_btn, "Exporta la lista de empleados a CSV")

        self.imprimir_btn = ttk.Button(btn_frame, text="Imprimir", command=self.imprimir_organigrama)
        self.imprimir_btn.grid(row=0, column=3, padx=5, pady=2)
        self.add_tooltip(self.imprimir_btn, "Imprime el organigrama")

        self.exportar_excel_btn = ttk.Button(btn_frame, text="Exportar plantilla Excel", command=self.exportar_plantilla_excel)
        self.exportar_excel_btn.grid(row=0, column=4, padx=5, pady=2)
        self.add_tooltip(self.exportar_excel_btn, "Genera una plantilla Excel para importar empleados")

        self.importar_excel_btn = ttk.Button(btn_frame, text="Importar empleados desde Excel", command=self.importar_empleados_excel)
        self.importar_excel_btn.grid(row=0, column=5, padx=5, pady=2)
        self.add_tooltip(self.importar_excel_btn, "Importa empleados desde un archivo Excel (.xlsx)")

        # Personalización de estilos
        estilo_frame = ttk.LabelFrame(self.root, text="Estilo del organigrama")
        estilo_frame.grid(row=7, column=0, padx=14, pady=5, sticky="ew")
        color_fondo_btn = ttk.Button(estilo_frame, text="Color de fondo", command=self.elegir_color_fondo)
        color_fondo_btn.grid(row=0, column=0, padx=2, pady=2)
        self.add_tooltip(color_fondo_btn, "Cambiar el color de fondo del organigrama")
        fuente_btn = ttk.Button(estilo_frame, text="Fuente", command=self.elegir_fuente)
        fuente_btn.grid(row=0, column=1, padx=2, pady=2)
        self.add_tooltip(fuente_btn, "Cambiar la fuente del organigrama")
        self.bordes_redondeados = tk.BooleanVar(value=True)
        bordes_chk = ttk.Checkbutton(estilo_frame, text="Bordes redondeados", variable=self.bordes_redondeados)
        bordes_chk.grid(row=0, column=2, padx=2, pady=2)
        self.add_tooltip(bordes_chk, "Activar o desactivar bordes redondeados en los cuadros")

        # Imagen del organigrama
        self.img_label = ttk.Label(self.root)
        self.img_label.grid(row=0, column=1, rowspan=8, padx=10, pady=10)

        # Estilo por defecto
        self.color_fondo = "#ffffff"
        self.fuente_actual = None
        self.fuente_nombre = "arial.ttf"
        self.fuente_tamano = 14

        # Arrastrar y soltar en la lista
        self.lista.bind('<Button-1>', self.inicio_arrastre)
        self.lista.bind('<B1-Motion>', self.arrastrando)
        self.lista.bind('<ButtonRelease-1>', self.soltar_arrastre)
        self._drag_data = {"indice": None}

    def elegir_color(self):
        color = colorchooser.askcolor()[1]
        if color:
            self.color_var.set(color)

    def agregar_empleado(self):
        nombre = self.nombre_var.get().strip()
        puesto = self.puesto_var.get().strip()
        depto = self.depto_var.get().strip()
        supervisor = self.supervisor_var.get().strip()
        color = self.color_var.get().strip()
        # Validaciones avanzadas
        if not nombre or not puesto or not depto:
            self.mostrar_notificacion("Nombre, Puesto y Departamento son obligatorios.", error=True)
            return
        if supervisor == nombre:
            self.mostrar_notificacion("Un empleado no puede ser su propio supervisor.", error=True)
            return
        # Prevenir ciclos en la jerarquía
        if self.crea_ciclo(nombre, supervisor):
            self.mostrar_notificacion("Asignar este supervisor generaría un ciclo en la jerarquía.", error=True)
            return
        # Verificar si ya existe (actualizar)
        for emp in self.empleados:
            if emp.nombre == nombre:
                emp.puesto = puesto
                emp.departamento = depto
                emp.supervisor = supervisor
                emp.color = color
                self.actualizar_lista()
                self.actualizar_supervisores()
                self.mostrar_notificacion("Empleado actualizado correctamente.")
                return
        # Nuevo
        self.empleados.append(Empleado(nombre, puesto, depto, supervisor, color))
        self.actualizar_lista()
        self.actualizar_supervisores()
        self.mostrar_notificacion("Empleado agregado correctamente.")

    def crea_ciclo(self, nombre, supervisor):
        # Checa si asignar 'supervisor' a 'nombre' genera un ciclo
        if not supervisor or supervisor == "(Ninguno)":
            return False
        cadena = set()
        actual = supervisor
        while actual and actual != "(Ninguno)":
            if actual == nombre:
                return True
            siguiente = None
            for emp in self.empleados:
                if emp.nombre == actual:
                    siguiente = emp.supervisor
                    break
            actual = siguiente
        return False

    def editar_empleado_rapido(self, event):
        seleccion = self.lista.curselection()
        if not seleccion:
            return
        idx = seleccion[0]
        nombre = self.lista.get(idx).split(" - ")[0]
        for emp in self.empleados:
            if emp.nombre == nombre:
                self.nombre_var.set(emp.nombre)
                self.puesto_var.set(emp.puesto)
                self.depto_var.set(emp.departamento)
                self.supervisor_var.set(emp.supervisor)
                self.color_var.set(emp.color)
                break

    def eliminar_empleado(self):
        seleccion = self.lista.curselection()
        if not seleccion:
            return
        idx = seleccion[0]
        nombre = self.lista.get(idx).split(" - ")[0]
        self.empleados = [emp for emp in self.empleados if emp.nombre != nombre]
        self.actualizar_lista()
        self.actualizar_supervisores()
        self.mostrar_notificacion("Empleado eliminado.")

    def seleccionar_empleado(self, event):
        seleccion = self.lista.curselection()
        if not seleccion:
            return
        idx = seleccion[0]
        nombre = self.lista.get(idx).split(" - ")[0]
        for emp in self.empleados:
            if emp.nombre == nombre:
                self.nombre_var.set(emp.nombre)
                self.puesto_var.set(emp.puesto)
                self.depto_var.set(emp.departamento)
                self.supervisor_var.set(emp.supervisor)
                self.color_var.set(emp.color)
                break

    def actualizar_lista(self):
        self.lista.delete(0, tk.END)
        for emp in self.empleados:
            self.lista.insert(tk.END, f"{emp.nombre} - {emp.puesto} - {emp.departamento}")
        # Resalta el seleccionado si existe
        seleccion = self.lista.curselection()
        if seleccion:
            self.lista.selection_set(seleccion[0])

    def actualizar_supervisores(self):
        nombres = [emp.nombre for emp in self.empleados]
        self.supervisor_combo["values"] = ["(Ninguno)"] + nombres

    def generar_organigrama(self):
        if not self.empleados:
            messagebox.showerror("Error", "No hay empleados para mostrar.")
            return
        # Construir diccionario de empleados por supervisor
        tree = {}
        root_nodes = []
        for emp in self.empleados:
            if not emp.supervisor or emp.supervisor == "(Ninguno)":
                root_nodes.append(emp)
            tree.setdefault(emp.supervisor, []).append(emp)
        # Parámetros de imagen y estilos
        ancho = 900
        alto = 600
        cuadro_w = 170
        cuadro_h = 60
        espacio_x = 30
        espacio_y = 80
        img = Image.new("RGB", (ancho, alto), self.color_fondo)
        draw = ImageDraw.Draw(img)
        # Fuente personalizada
        try:
            font = ImageFont.truetype(self.fuente_nombre, self.fuente_tamano)
        except:
            font = ImageFont.load_default()
        # Layout recursivo
        posiciones = {}
        def layout(node, x, y):
            posiciones[node.nombre] = (x, y)
            hijos = tree.get(node.nombre, [])
            if hijos:
                total_w = (len(hijos)-1)*(cuadro_w+espacio_x)
                start_x = x - total_w//2
                for i, hijo in enumerate(hijos):
                    hx = start_x + i*(cuadro_w+espacio_x)
                    hy = y + cuadro_h + espacio_y
                    layout(hijo, hx, hy)
        # Distribuir raíces
        if len(root_nodes) == 0:
            messagebox.showerror("Error", "No hay nodo raíz (sin supervisor).")
            return
        total_w = (len(root_nodes)-1)*(cuadro_w+espacio_x)
        start_x = ancho//2 - total_w//2
        for i, root in enumerate(root_nodes):
            rx = start_x + i*(cuadro_w+espacio_x)
            ry = 40
            layout(root, rx, ry)
        # Dibujar líneas
        for emp in self.empleados:
            if emp.supervisor and emp.supervisor != "(Ninguno)":
                if emp.supervisor in posiciones:
                    x1, y1 = posiciones[emp.supervisor]
                    x2, y2 = posiciones[emp.nombre]
                    draw.line([(x1+cuadro_w//2, y1+cuadro_h), (x2+cuadro_w//2, y2)], fill="black", width=2)
        # Dibujar cuadros
        for emp in self.empleados:
            x, y = posiciones[emp.nombre]
            if self.bordes_redondeados.get():
                radio = 18
                draw.rounded_rectangle([x, y, x+cuadro_w, y+cuadro_h], radius=radio, fill=emp.color, outline="black", width=2)
            else:
                draw.rectangle([x, y, x+cuadro_w, y+cuadro_h], fill=emp.color, outline="black", width=2)
            texto = f"{emp.nombre}\n{emp.puesto}\n{emp.departamento}"
            lines = texto.split("\n")
            for i, line in enumerate(lines):
                bbox = draw.textbbox((0, 0), line, font=font)
                w = bbox[2] - bbox[0]
                h = bbox[3] - bbox[1]
                draw.text((x + (cuadro_w-w)//2, y + 10 + i*18), line, fill="black", font=font)
        # Mostrar imagen
        img = img.resize((600, 400), Image.LANCZOS)
        self.org_img = ImageTk.PhotoImage(img)
        self.img_label.configure(image=self.org_img)
        self.img_label.image = self.org_img
        self.org_img_raw = img.copy()  # Guardar la imagen PIL para exportar
        self.exportar_png_btn.config(state="normal")
        self.exportar_pdf_btn.config(state="normal")
        self.mostrar_notificacion("Organigrama generado.")

    def exportar_png(self):
        if not hasattr(self, 'org_img_raw'):
            self.mostrar_notificacion("Primero genera el organigrama.", error=True)
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG files", "*.png")])
        if file_path:
            self.org_img_raw.save(file_path, "PNG")
            self.mostrar_notificacion(f"Organigrama guardado como PNG en\n{file_path}")

    def exportar_pdf(self):
        if not hasattr(self, 'org_img_raw'):
            self.mostrar_notificacion("Primero genera el organigrama.", error=True)
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if file_path:
            # Guardar imagen temporal
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                temp_img_path = tmp.name
                self.org_img_raw.save(temp_img_path, "PNG")
            c = canvas.Canvas(file_path, pagesize=letter)
            width, height = letter
            # Ajustar tamaño de imagen al ancho de la página
            img_width, img_height = self.org_img_raw.size
            aspect = img_height / img_width
            pdf_width = width - 80
            pdf_height = pdf_width * aspect
            c.drawImage(temp_img_path, 40, height - 40 - pdf_height, width=pdf_width, height=pdf_height)
            c.save()
            os.remove(temp_img_path)
            self.mostrar_notificacion(f"Organigrama guardado como PDF en\n{file_path}")

    def guardar_empleados(self):
        if not self.empleados:
            messagebox.showerror("Error", "No hay empleados para guardar.")
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("Archivos JSON", "*.json")], initialfile="empleados.json")
        if file_path:
            data = [
                {
                    "nombre": emp.nombre,
                    "puesto": emp.puesto,
                    "departamento": emp.departamento,
                    "supervisor": emp.supervisor,
                    "color": emp.color
                }
                for emp in self.empleados
            ]
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            messagebox.showinfo("Éxito", f"Empleados guardados en\n{file_path}")

    def cargar_empleados(self):
        file_path = filedialog.askopenfilename(defaultextension=".json", filetypes=[("Archivos JSON", "*.json")])
        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                self.empleados = [Empleado(emp["nombre"], emp["puesto"], emp["departamento"], emp["supervisor"], emp["color"]) for emp in data]
                self.actualizar_lista()
                self.actualizar_supervisores()
                self.mostrar_notificacion(f"Empleados cargados desde\n{file_path}")
            except Exception as e:
                self.mostrar_notificacion(f"No se pudo cargar el archivo:\n{e}", error=True)

    def exportar_csv(self):
        if not self.empleados:
            self.mostrar_notificacion("No hay empleados para exportar.", error=True)
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("Archivos CSV", "*.csv")])
        if file_path:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Nombre", "Puesto", "Departamento", "Supervisor", "Color"])
                for emp in self.empleados:
                    writer.writerow([emp.nombre, emp.puesto, emp.departamento, emp.supervisor, emp.color])
            self.mostrar_notificacion(f"Empleados exportados a CSV en\n{file_path}")

    def imprimir_organigrama(self):
        if not hasattr(self, 'org_img_raw'):
            self.mostrar_notificacion("Primero genera el organigrama.", error=True)
            return
        # Guardar imagen temporal
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
            temp_img_path = tmp.name
            self.org_img_raw.save(temp_img_path, "PNG")
        # Abrir diálogo de impresión del sistema
        if platform.system() == "Windows":
            os.startfile(temp_img_path, "print")
        else:
            self.mostrar_notificacion(f"Abre el archivo {temp_img_path} y usa Ctrl+P para imprimir.")

    def elegir_color_fondo(self):
        color = colorchooser.askcolor(title="Color de fondo")[1]
        if color:
            self.color_fondo = color

    def elegir_fuente(self):
        fuente = simpledialog.askstring("Fuente", "Nombre del archivo de fuente (.ttf):", initialvalue=self.fuente_nombre)
        if fuente:
            self.fuente_nombre = fuente
        tamano = simpledialog.askinteger("Tamaño de fuente", "Tamaño de fuente:", initialvalue=self.fuente_tamano)
        if tamano:
            self.fuente_tamano = tamano

    def exportar_plantilla_excel(self):
        if openpyxl is None:
            messagebox.showerror("Error", "La librería openpyxl no está instalada. Instálala con:\npip install openpyxl")
            return
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Empleados"
        encabezados = ["Nombre", "Puesto", "Departamento", "Supervisor", "Color"]
        ws.append(encabezados)
        ejemplo = ["Juan Pérez", "Gerente", "Ventas", "(Ninguno)", "#A3C9F1"]
        ws.append(ejemplo)
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
        ruta = os.path.join(os.path.dirname(os.path.abspath(__file__)), "plantilla_empleados.xlsx")
        try:
            wb.save(ruta)
            self.mostrar_notificacion("Plantilla Excel generada: plantilla_empleados.xlsx", color="#1976d2")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar la plantilla: {e}")

    def importar_empleados_excel(self):
        if openpyxl is None:
            messagebox.showerror("Error", "La librería openpyxl no está instalada. Instálala con:\npip install openpyxl")
            return
        ruta = filedialog.askopenfilename(title="Selecciona el archivo Excel", filetypes=[("Archivos Excel", "*.xlsx")])
        if not ruta:
            return
        from openpyxl import load_workbook
        try:
            wb = load_workbook(ruta)
            ws = wb.active
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el archivo: {e}")
            return
        encabezados = ["Nombre", "Puesto", "Departamento", "Supervisor", "Color"]
        fila_encabezados = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if fila_encabezados != encabezados:
            messagebox.showerror("Error", "El archivo Excel no tiene los encabezados correctos. Usa la plantilla generada por el programa.")
            return
        errores = []
        nuevos = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            nombre, puesto, depto, supervisor, color = row
            if not nombre or not puesto or not depto:
                errores.append(f"Fila {i}: Faltan campos obligatorios.")
                continue
            if not color or not str(color).startswith("#") or len(str(color)) not in (7, 9):
                errores.append(f"Fila {i}: Color inválido.")
                continue
            nuevos.append((nombre, puesto, depto, supervisor, color))
        if errores:
            messagebox.showerror("Errores al importar", "\n".join(errores))
        for nombre, puesto, depto, supervisor, color in nuevos:
            self.empleados.append(Empleado(nombre, puesto, depto, supervisor, color))
        if nuevos:
            self.actualizar_lista()
            self.actualizar_supervisores()
            self.mostrar_notificacion(f"{len(nuevos)} empleados importados correctamente.", color="#388e3c")

    # --- Arrastrar y soltar en la lista para cambiar supervisor ---
    def inicio_arrastre(self, event):
        idx = self.lista.nearest(event.y)
        if idx >= 0:
            self._drag_data["indice"] = idx

    def arrastrando(self, event):
        pass  # Solo visual

    def soltar_arrastre(self, event):
        if self._drag_data["indice"] is None:
            return
        idx_origen = self._drag_data["indice"]
        idx_destino = self.lista.nearest(event.y)
        if idx_origen == idx_destino or idx_destino < 0:
            self._drag_data["indice"] = None
            return
        nombre_origen = self.lista.get(idx_origen).split(" - ")[0]
        nombre_destino = self.lista.get(idx_destino).split(" - ")[0]
        if nombre_origen == nombre_destino:
            self._drag_data["indice"] = None
            return
        # Validar que no se genere ciclo
        if self.crea_ciclo(nombre_origen, nombre_destino):
            messagebox.showerror("Error", "No se puede asignar este supervisor porque generaría un ciclo.")
            self._drag_data["indice"] = None
            return
        # Cambiar supervisor
        for emp in self.empleados:
            if emp.nombre == nombre_origen:
                emp.supervisor = nombre_destino
                break
        self.actualizar_lista()
        self.actualizar_supervisores()

def on_close():
    # Autoguardado silencioso
    ruta = os.path.join(os.path.dirname(os.path.abspath(__file__)), "empleados.json")
    data = [
        {
            "nombre": emp.nombre,
            "puesto": emp.puesto,
            "departamento": emp.departamento,
            "supervisor": emp.supervisor,
            "color": emp.color
        }
        for emp in app.empleados
    ]
    try:
        with open(ruta, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except:
        pass
    root.destroy()

root = tk.Tk()
app = OrgApp(root)
root.protocol("WM_DELETE_WINDOW", on_close)
root.mainloop()
